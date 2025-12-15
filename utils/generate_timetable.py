import pandas as pd
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import re

# --- Load Data ---
def load_data():
    timetable_df = pd.read_csv('Data/final_timetable.csv')
    courses_df = pd.read_csv('Data/Courses.csv').apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    sections_input = {
        "CSIT": {
            "1": [{"name": f"s{i}", "count": 25} for i in range(1, 10)],
            "2": [{"name": f"s{i}", "count": 25} for i in range(1, 10)],
            "3": {
                "AID": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}, {"name": "s3", "count": 25}],
                "CNC": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}, {"name": "s3", "count": 25}],
                "CSC": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}],
                "BIF": [{"name": "s1", "count": 25}]
            },
            "4": {
                "AID": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}, {"name": "s3", "count": 25}],
                "CNC": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}, {"name": "s3", "count": 25}],
                "CSC": [{"name": "s1", "count": 25}, {"name": "s2", "count": 25}],
                "BIF": [{"name": "s1", "count": 25}]
            }
        }
    }
    return timetable_df, sections_input, courses_df

# --- Generate Excel Timetable ---
def generate_excel(timetable_df, level, specialization, sections, sections_input, courses_df):
    # Parse level and specialization
    level_num = level.split('-')[0] if '-' in level else level
    spec = specialization if specialization else None

    # If specialization is None and level is 3 or 4, combine all specializations
    if spec is None and level_num in ['3', '4']:
        all_specs = sections_input['CSIT'][level_num]
        sections_list = []
        spec_sections = {}
        for s, secs in all_specs.items():
            spec_secs = [f"CSIT-{level_num}-{s}-s{i+1}" for i in range(len(secs))]
            sections_list.extend(spec_secs)
            spec_sections[s] = len(secs)
        combined = True
        title = f"CSIT {int(level_num)}rd Year Timetable"
    else:
        combined = False
        if spec:
            sections_list = [f"CSIT-{level_num}-{spec}-s{i+1}" for i in range(len(sections))]
            spec_sections = {spec: len(sections)}
        else:
            sections_list = [f"CSIT-{level_num}-s{i+1}" for i in range(len(sections))]
            spec_sections = {'Core': len(sections)}
        title = f"CSIT Level {level}{'-' + spec if spec else ''} Timetable"

    # Filter timetable for the specified level and specialization(s)
    section_pattern = '|'.join(sections_list)
    filtered_df = timetable_df[timetable_df['Sections'].str.contains(section_pattern, na=False)]

    # Define timeslots and days
    timeslots = ['9:00 AM', '10:45 AM', '12:30 PM', '2:15 PM']
    timeslot_ranges = ['9:00 AM\nto 10:30 AM', '10:45 AM\nto 12:15 PM', '12:30 PM\n to 2:00 PM', '2:15 PM\nto 3:45 PM']
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
    timeslot_order = {t: i for i, t in enumerate(timeslots)}

    # Sort timetable by day and start time
    filtered_df = filtered_df.sort_values(by=['Day', 'StartTime'], key=lambda x: x.map(lambda y: days.index(y) if x.name == 'Day' else timeslot_order.get(y, len(timeslots))))

    # Create course name lookup
    course_names = dict(zip(courses_df['CourseID'], courses_df['CourseName']))

    # Unique CourseIDs and assign colors
    unique_course_ids = filtered_df['CourseID'].unique()
    # 20 light colors suitable for black text
    colors = [
  "FFB300",
  "FF6F00",
  "FF4081",
  "D500F9",
  "651FFF",
  "2962FF",
  "00B0FF",
  "00E5FF",
  "1DE9B6",
  "76FF03",
  "C6FF00",
  "FFEA00",
  "FFAB00",
  "FF7043",
  "A1887F",
  "8D6E63",
  "6D4C41",
  "546E7A",
  "37474F",
  "B0BEC5"
  ]
    course_colors = {cid: colors[i % len(colors)] for i, cid in enumerate(unique_course_ids)}

    # Create timetable dictionary
    timetable_dict = defaultdict(lambda: defaultdict(tuple))
    lecture_spans = defaultdict(list)
    for _, row in filtered_df.iterrows():
        day = row['Day']
        time = row['StartTime']
        sections = row['Sections'].split(', ')
        course_name = course_names.get(row['CourseID'], '')
        course_id_formatted = re.sub(r'([A-Z]+)(\d+)', r'\1 \2', row['CourseID'])
        type_abbr = row['Type'][:3].upper()
        cell_content = f"{course_id_formatted} {course_name}\n{row['Instructor']}\n{type_abbr} {row['Room']}"
        color = course_colors.get(row['CourseID'], 'FFFFFF')
        relevant_sections = [s for s in sections if s in sections_list]
        if relevant_sections:
            if row['Type'] == 'Lecture':
                # Store lecture to handle merging, now including color
                lecture_spans[(day, time, cell_content)] = ([s.split('-')[-1] for s in relevant_sections], color)
            else:
                # For labs, assign to individual sections
                for section in relevant_sections:
                    section_key = section.split('-')[-1]
                    timetable_dict[(day, time)][section_key] = (cell_content, color)

    # Add merged lectures
    # *** FIXED HERE: Unpack the (section_keys, color) tuple ***
    for (day, time, cell_content), (section_keys, color) in lecture_spans.items():
        for section_key in section_keys:
            timetable_dict[(day, time)][section_key] = (cell_content, color)

    # Initialize Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title.replace(' ', '_')

    # Write headers
    ws['A1'] = title
    ws.merge_cells('A1:' + get_column_letter(len(sections_list) + 2) + '1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Write column headers
    ws['A2'] = "Day"
    ws['B2'] = "Time"
    col_idx = 3
    for s, num in spec_sections.items():
        start_col = col_idx
        ws.cell(row=2, column=start_col).value = f"CSIT  {s}" if combined else s
        ws.cell(row=2, column=start_col).font = Font(bold=True)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=start_col).fill = PatternFill(start_color='ffdab9', end_color='ffdab9', fill_type='solid')
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + num - 1)
        for j in range(num):
            ws.cell(row=3, column=col_idx).value = f"s{j+1}"
            ws.cell(row=3, column=col_idx).font = Font(bold=True)
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal='center')
            col_idx += 1

    # Write day and time slots, merging days
    row_idx = 4
    for day in days:
        day_row_start = row_idx
        for i, _ in enumerate(timeslots):
            ws.cell(row=row_idx, column=1).value = day
            ws.cell(row=row_idx, column=2).value = timeslot_ranges[i]
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            row_idx += 1
        # Merge day cells across the 4 time slots
        ws.merge_cells(start_row=day_row_start, start_column=1, end_row=row_idx-1, end_column=1)
        ws.cell(row=day_row_start, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Fill timetable
    row_idx = 4
    for day in days:
        for time in timeslots:
            col = 3
            while col < len(sections_list) + 3:
                section = sections_list[col-3]
                section_key = section.split('-')[-1]
                data = timetable_dict[(day, time)].get(section_key, ('', 'FFFFFF'))
                cell_content, fill_color = data
                if cell_content:
                    ws.cell(row=row_idx, column=col).value = cell_content
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row=row_idx, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                # Merge cells for same content covering multiple sections
                if cell_content:
                    start_col = col
                    end_col = col
                    while end_col < len(sections_list) + 3 and timetable_dict[(day, time)].get(sections_list[end_col-3].split('-')[-1], ('', ''))[0] == cell_content:
                        end_col += 1
                    if end_col > start_col:
                        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col - 1)
                        ws.cell(row=row_idx, column=start_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    col = end_col
                else:
                    col += 1
            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    for col in range(3, len(sections_list) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 20


    # Save Excel file
    output_dir = "timetables_output/"  # Relative to current directory
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"timetable_level_{level}{'-' + spec if spec else ''}.xlsx")
    wb.save(output_file)
    print(f"✅ Excel timetable saved to '{output_file}'")
    
    print("To export to PDF, open the Excel file and save as PDF, or open the HTML in a browser and print to PDF.")

# --- Generate HTML Timetable ---
def generate_html(timetable_df, level, specialization, sections, sections_input, courses_df):
    # Parse level and specialization
    level_num = level.split('-')[0] if '-' in level else level
    spec = specialization if specialization else None

    # If specialization is None and level is 3 or 4, combine all specializations
    if spec is None and level_num in ['3', '4']:
        all_specs = sections_input['CSIT'][level_num]
        sections_list = []
        spec_sections = {}
        for s, secs in all_specs.items():
            spec_secs = [f"CSIT-{level_num}-{s}-s{i+1}" for i in range(len(secs))]
            sections_list.extend(spec_secs)
            spec_sections[s] = len(secs)
        combined = True
        title = f"CSIT {int(level_num)}rd Year Timetable"
    else:
        combined = False
        if spec:
            sections_list = [f"CSIT-{level_num}-{spec}-s{i+1}" for i in range(len(sections))]
            spec_sections = {spec: len(sections)}
        else:
            sections_list = [f"CSIT-{level_num}-s{i+1}" for i in range(len(sections))]
            spec_sections = {'Core': len(sections)}
        title = f"CSIT Level {level}{'-' + spec if spec else ''} Timetable"

    # Filter timetable for the specified level and specialization(s)
    section_pattern = '|'.join(sections_list)
    filtered_df = timetable_df[timetable_df['Sections'].str.contains(section_pattern, na=False)]

    # Define timeslots and days
    timeslots = ['9:00 AM', '10:45 AM', '12:30 PM', '2:15 PM']
    timeslot_ranges = ['9:00 AM\nto 10:30 AM', '10:45 AM\nto 12:15 PM', '12:30 PM\nto 2:00 PM', '2:15 PM\nto 3:45 PM']
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
    timeslot_order = {t: i for i, t in enumerate(timeslots)}

    # Sort timetable by day and start time
    filtered_df = filtered_df.sort_values(by=['Day', 'StartTime'], key=lambda x: x.map(lambda y: days.index(y) if x.name == 'Day' else timeslot_order.get(y, len(timeslots))))

    # Create course name lookup
    course_names = dict(zip(courses_df['CourseID'], courses_df['CourseName']))

    # Unique CourseIDs and assign colors
    unique_course_ids = filtered_df['CourseID'].unique()
    # 20 light colors suitable for black text
    colors = [
  "FFB300",
  "FF6F00",
  "FF4081",
  "D500F9",
  "651FFF",
  "2962FF",
  "00B0FF",
  "00E5FF",
  "1DE9B6",
  "76FF03",
  "C6FF00",
  "FFEA00",
  "FFAB00",
  "FF7043",
  "A1887F",
  "8D6E63",
  "6D4C41",
  "546E7A",
  "37474F",
  "B0BEC5"
]

    course_colors = {cid: colors[i % len(colors)] for i, cid in enumerate(unique_course_ids)}

    # Create timetable dictionary
    timetable_dict = defaultdict(lambda: defaultdict(tuple))
    lecture_spans = defaultdict(list)
    for _, row in filtered_df.iterrows():
        day = row['Day']
        time = row['StartTime']
        sections = row['Sections'].split(', ')
        course_name = course_names.get(row['CourseID'], '')
        course_id_formatted = re.sub(r'([A-Z]+)(\d+)', r'\1 \2', row['CourseID'])
        type_abbr = row['Type'][:3].upper()
        cell_content = f"{course_id_formatted} {course_name}\n{row['Instructor']}\n{type_abbr} {row['Room']}"
        color = course_colors.get(row['CourseID'], 'FFFFFF')
        relevant_sections = [s for s in sections if s in sections_list]
        if relevant_sections:
            if row['Type'] == 'Lecture':
                # Store lecture to handle merging, now including color
                lecture_spans[(day, time, cell_content)] = ([s.split('-')[-1] for s in relevant_sections], color)
            else:
                # For labs, assign to individual sections
                for section in relevant_sections:
                    section_key = section.split('-')[-1]
                    timetable_dict[(day, time)][section_key] = (cell_content, color)

    # Add merged lectures
    # *** FIXED HERE: Unpack the (section_keys, color) tuple ***
    for (day, time, cell_content), (section_keys, color) in lecture_spans.items():
        for section_key in section_keys:
            timetable_dict[(day, time)][section_key] = (cell_content, color)

    # Generate HTML
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{title}</title>
        <style>
            table {{ border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; }}
            th, td {{ border: 1px solid black; padding: 8px; text-align: center; }}
            th {{ background-color: #f2f2f2; }}
            h1 {{ text-align: center; font-family: Arial, sans-serif; }}
            td {{ white-space: pre-line; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <table>
            <tr>
                <th rowspan="2">Day</th>
                <th rowspan="2">Time</th>
                {"".join(f'<th colspan="{num}" style="background-color: #ffdab9;">CSIT  {s}</th>' for s, num in spec_sections.items())}
            </tr>
            <tr>
                {"".join("".join(f'<th>s{j+1}</th>' for j in range(num)) for s, num in spec_sections.items())}
            </tr>
    """

    for day in days:
        html_content += f'<tr><td rowspan="4">{day}</td>'
        for i, time in enumerate(timeslots):
            if i != 0:  # Add new row for subsequent times
                html_content += '<tr>'
            html_content += f'<td>{timeslot_ranges[i]}</td>'
            used_sections = set()  # Track sections already covered by merged lectures
            for section in sections_list:
                section_key = section.split('-')[-1]
                data = timetable_dict[(day, time)].get(section_key, ('', 'FFFFFF'))
                cell_content, color = data
                
                if section_key in used_sections:
                    continue
                
                # *** FIXED HERE: Correctly check the tuple in lecture_spans ***
                lecture_span_data = lecture_spans.get((day, time, cell_content))
                
                if cell_content and 'LEC' in cell_content and lecture_span_data and section_key in lecture_span_data[0]:
                    # Count how many sections this lecture spans
                    span_sections = lecture_span_data[0] # Get the section list (index 0)
                    if len(span_sections) > 1:
                        html_content += f'<td colspan="{len(span_sections)}" style="background-color: #{color};">{cell_content}</td>'
                        used_sections.update(span_sections)
                    else:
                        html_content += f'<td style="background-color: #{color};">{cell_content}</td>'
                        used_sections.add(section_key) # Add even single lectures to avoid re-processing
                elif cell_content:
                    html_content += f'<td style="background-color: #{color};">{cell_content}</td>'
                else:
                    html_content += '<td></td>'
            html_content += '</tr>'
    html_content += """
        </table>
    </body>
    </html>
    """

    # Save HTML file
    output_dir = "timetables_output/"  # Relative to current directory
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"timetable_level_{level}{'-' + spec if spec else ''}.html")
    with open(output_file, 'w') as f:
        f.write(html_content)
    print(f"✅ HTML timetable saved to '{output_file}'")

# --- Main Function ---
def main():
    try:
        timetable_df, sections_input, courses_df = load_data()
    except FileNotFoundError as e:
        print(f"Error: Could not find a required data file.")
        print(f"Missing file: {e.filename}")
        print("Please make sure 'final_timetable_first.csv' and 'CSIT_Courses.csv' are in the same directory.")
        return
    except Exception as e:
        print(f"An error occurred while loading data: {e}")
        return

    # Get available levels
    levels = []
    for level, level_data in sections_input['CSIT'].items():
        if isinstance(level_data, list):
            levels.append(str(level))
        else:
            for spec in level_data.keys():
                levels.append(f"{level}-{spec}")
    print("Available levels:", ', '.join(levels))

    # Prompt for level
    level_input = input("Enter the level (e.g., '1', '3-AID', '4-CSC'): ").strip()
    if level_input not in levels:
        print(f"Error: '{level_input}' is not a valid level. Please choose from: {', '.join(levels)}")
        return

    # Parse level and specialization
    if '-' in level_input:
        level, specialization = level_input.split('-')
        sections = sections_input['CSIT'][level][specialization]
    else:
        level = level_input
        specialization = None
        sections = sections_input['CSIT'].get(level, [])

    # Generate outputs
    generate_excel(timetable_df, level, specialization, sections, sections_input, courses_df)
    generate_html(timetable_df, level, specialization, sections, sections_input, courses_df)

if __name__ == "__main__":
    main()