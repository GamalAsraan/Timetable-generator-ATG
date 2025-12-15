import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import re
import webbrowser
import os

def load_data():
    """Loads the required CSV files."""
    try:
        timetable_df = pd.read_csv('Data/final_timetable.csv')
        courses_df = pd.read_csv('Data/Courses.csv').apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    except FileNotFoundError as e:
        print(f"Error: Could not find a required data file.")
        print(f"Missing file: {e.filename}")
        print("Please make sure 'final_timetable_first.csv' and 'CSIT_Courses.csv' are in the same directory.")
        return None, None
    except Exception as e:
        print(f"An error occurred while loading data: {e}")
        return None, None
    return timetable_df, courses_df

def generate_instructor_timetable(instructor_name, timetable_df, courses_df):
    """Generates HTML and Excel timetables for a specific instructor."""
    
    # Filter for the instructor
    instructor_df = timetable_df[timetable_df['Instructor'] == instructor_name].copy()

    if instructor_df.empty:
        print(f"No schedule found for '{instructor_name}'.")
        return

    # Define timeslots and days
    timeslots = ['9:00 AM', '10:45 AM', '12:30 PM', '2:15 PM']
    timeslot_ranges = ['9:00 AM\nto 10:30 AM', '10:45 AM\nto 12:15 PM', '12:30 PM\nto 2:00 PM', '2:15 PM\nto 3:45 PM']
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']

    # Create course name lookup
    course_names = dict(zip(courses_df['CourseID'], courses_df['CourseName']))

    # Create a simple schedule dictionary: schedule[(day, time)] = cell_content
    schedule = defaultdict(str)
    
    # Sort entries by day and time for consistent processing
    timeslot_order = {t: i for i, t in enumerate(timeslots)}
    day_order = {d: i for i, d in enumerate(days)}
    
    instructor_df['day_sort'] = instructor_df['Day'].map(day_order)
    instructor_df['time_sort'] = instructor_df['StartTime'].map(timeslot_order)
    instructor_df = instructor_df.sort_values(by=['day_sort', 'time_sort'])

    for _, row in instructor_df.iterrows():
        day = row['Day']
        time = row['StartTime']
        
        # Format cell content
        course_id_formatted = re.sub(r'([A-Z]+)(\d+)', r'\1 \2', row['CourseID'])
        course_name = course_names.get(row['CourseID'], '')
        type_abbr = row['Type'][:3].upper()
        
        # Clean up section list for display
        sections = row['Sections'].replace(' ', '') # Remove spaces
        
        cell_content = (
            f"{course_id_formatted} {course_name}\n"
            f"{type_abbr} ({sections})\n"
            f"Room: {row['Room']}"
        )
        
        # Handle multiple entries in the same slot (e.g., different labs)
        if (day, time) in schedule:
            schedule[(day, time)] += f"\n---\n{cell_content}"
        else:
            schedule[(day, time)] = cell_content

    # --- Generate Excel File ---
    wb = Workbook()
    ws = wb.active
    safe_name = re.sub(r'[^a-zA-Z0-9]', '_', instructor_name)
    ws.title = safe_name
    title = f"Timetable for {instructor_name}"

    # Title
    ws['A1'] = title
    ws.merge_cells(f'A1:{get_column_letter(len(timeslots) + 1)}1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    ws['A2'] = "Day"
    ws['A2'].font = Font(bold=True)
    for i, time_range in enumerate(timeslot_ranges, start=2):
        col = get_column_letter(i)
        ws[f'{col}2'] = time_range.replace("\n", " ")
        ws[f'{col}2'].font = Font(bold=True)
        ws[f'{col}2'].alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[col].width = 30

    ws.column_dimensions['A'].width = 15

    # Fill schedule
    row_idx = 3
    for day in days:
        ws.cell(row=row_idx, column=1).value = day
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = Alignment(vertical='center', horizontal='center')
        ws.row_dimensions[row_idx].height = 100 # Set row height
        
        for col_idx, time in enumerate(timeslots, start=2):
            content = schedule.get((day, time), '')
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = content
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if content:
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        row_idx += 1

    excel_file = f"timetable_{safe_name}.xlsx"
    wb.save(excel_file)
    print(f"✅ Excel timetable saved to '{excel_file}'")

    # --- Generate HTML File ---
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{title}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ text-align: center; }}
            table {{ 
                border-collapse: collapse; 
                width: 90%; 
                margin: 20px auto; 
                box-shadow: 0 4px 10px rgba(0,0,0,0.1);
            }}
            th, td {{ 
                border: 1px solid #ddd; 
                padding: 12px; 
                text-align: center; 
                min-height: 100px;
                vertical-align: top;
            }}
            th {{ 
                background-color: #004a99; 
                color: white; 
                font-weight: bold;
            }}
            td {{ 
                white-space: pre-line; /* Respects newlines */
                background-color: #f9f9f9;
            }}
            tr:nth-child(even) td {{ background-color: #f2f2f2; }}
            td:first-child {{ 
                font-weight: bold; 
                background-color: #f0f0f0;
                width: 10%;
            }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <table>
            <tr>
                <th>Day / Time</th>
    """
    # HTML Headers
    for time_range in timeslot_ranges:
        html_content += f"<th>{time_range.replace(chr(10), '<br>')}</th>"
    html_content += "</tr>"

    # HTML Body
    for day in days:
        html_content += f"<tr><td>{day}</td>"
        for time in timeslots:
            content = schedule.get((day, time), '')
            # Convert newlines to <br> and --- to <hr> for HTML
            content_html = content.replace("\n---\n", "<hr style='margin: 8px 0;'>").replace("\n", "<br>")
            
            if content_html:
                html_content += f'<td style="background-color: #e6f7ff;">{content_html}</td>'
            else:
                html_content += '<td></td>'
        html_content += "</tr>"

    html_content += """
        </table>
    </body>
    </html>
    """

    html_file = f"timetable_{safe_name}.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"✅ HTML timetable saved to '{html_file}'")
    
    # Open the HTML file in the default browser
    try:
        webbrowser.open('file://' + os.path.realpath(html_file))
    except Exception as e:
        print(f"Could not open HTML file in browser: {e}")


def main():
    timetable_df, courses_df = load_data()
    if timetable_df is None:
        return

    # Get and display unique instructor names
    instructors = sorted(timetable_df['Instructor'].unique())
    print("--- Available Instructors ---")
    for i, name in enumerate(instructors):
        print(f"  {name}")
    print("-----------------------------")

    # Prompt user for instructor name
    instructor_name = input("Enter the full name of the instructor to generate a timetable for: ").strip()

    if instructor_name not in instructors:
        print(f"\nError: Instructor '{instructor_name}' not found.")
        print("Please make sure the name is spelled exactly as shown in the list.")
    else:
        print(f"\nGenerating timetable for {instructor_name}...")
        generate_instructor_timetable(instructor_name, timetable_df, courses_df)

if __name__ == "__main__":
    main()